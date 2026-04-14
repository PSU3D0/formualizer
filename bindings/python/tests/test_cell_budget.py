from pathlib import Path

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


def _dense_xlsx(path: Path, rows: int, cols: int) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c, value=1)
    wb.save(path)


def test_recalculate_file_cell_budget_rejects_over_budget(tmp_path: Path):
    path = tmp_path / "dense.xlsx"
    _dense_xlsx(path, rows=11, cols=10)

    with pytest.raises(IOError) as excinfo:
        fz.recalculate_file(str(path), cell_budget=50)
    assert "logical-cell budget" in str(excinfo.value)


def test_recalculate_file_cell_budget_generous_succeeds(tmp_path: Path):
    in_path = tmp_path / "in.xlsx"
    out_path = tmp_path / "out.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 2
    ws["A2"] = 3
    ws["B1"] = "=A1+A2"
    wb.save(in_path)

    summary = fz.recalculate_file(
        str(in_path), output=str(out_path), cell_budget=10_000_000
    )
    assert summary["status"] == "success"
    assert summary["evaluated"] == 1


def test_workbook_config_cell_budget_rejects_on_load(tmp_path: Path):
    path = tmp_path / "dense.xlsx"
    _dense_xlsx(path, rows=11, cols=10)

    cfg = fz.WorkbookConfig(cell_budget=50)
    with pytest.raises(Exception) as excinfo:
        fz.load_workbook(str(path), config=cfg)
    assert "logical-cell budget" in str(excinfo.value)


def test_workbook_config_cell_budget_generous_loads(tmp_path: Path):
    path = tmp_path / "small.xlsx"
    _dense_xlsx(path, rows=3, cols=3)

    cfg = fz.WorkbookConfig(cell_budget=10_000_000)
    wb = fz.load_workbook(str(path), config=cfg)
    assert wb is not None

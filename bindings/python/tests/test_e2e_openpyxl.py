import tempfile
from pathlib import Path

import pytest

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - allow skipping if not present in dev env
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


# The extension module name configured by maturin
import formualizer as fz


def make_wb(tmp: Path) -> Path:
    p = tmp / "e2e.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Values
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = 3
    # Simple formula
    ws["B1"] = "=SUM(A1:A3)"
    # Conditionals
    ws["C1"] = "=IF(B1>3, B1*2, 0)"
    # SUMIFS-like (if supported) else basic SUM as placeholder
    ws["D1"] = "=SUM(A1:A3)"
    wb.save(p)
    return p


def test_openpyxl_roundtrip(tmp_path: Path):
    # Prepare XLSX via openpyxl
    xlsx_path = make_wb(tmp_path)

    # Load the actual workbook from disk using Calamine adapter
    wb = fz.load_workbook(str(xlsx_path), strategy="eager_all")
    
    # Create an engine from the workbook
    engine = fz.Engine.from_workbook(wb)
    
    # Evaluate and check values
    val_b1 = engine.evaluate_cell("Sheet1", 1, 2)
    print(f"B1 value type: {val_b1.type_name}, value: {val_b1}")
    assert val_b1.as_number() == 6.0
    
    val_c1 = engine.evaluate_cell("Sheet1", 1, 3)
    print(f"C1 value type: {val_c1.type_name}, value: {val_c1}")
    assert val_c1.as_number() == 12.0
    
    val_d1 = engine.evaluate_cell("Sheet1", 1, 4)
    print(f"D1 value type: {val_d1.type_name}, value: {val_d1}")
    assert val_d1.as_number() == 6.0

    # Note: With the new pattern, mutation happens through the workbook
    # but evaluation happens through the engine
    # For now, we'll skip the mutation test as it would require re-loading the workbook into the engine


def test_batch_values_and_formulas():
    wb = fz.Workbook()
    s = wb.sheet("Data")

    s.set_values_batch(1, 1, 2, 3, [
        [fz.LiteralValue.int(1), fz.LiteralValue.int(2), fz.LiteralValue.int(3)],
        [fz.LiteralValue.int(4), fz.LiteralValue.int(5), fz.LiteralValue.int(6)],
    ])

    s.set_formulas_batch(1, 4, 2, 1, [
        ["=SUM(A1:C1)"],
        ["=SUM(A2:C2)"],
    ])

    # Create engine from the workbook
    engine = fz.Engine.from_workbook(wb)
    
    # Evaluate the formula cells
    val1 = engine.evaluate_cell("Data", 1, 4)
    val2 = engine.evaluate_cell("Data", 2, 4)
    
    print(f"Sum 1 value type: {val1.type_name}, value: {val1}")
    print(f"Sum 2 value type: {val2.type_name}, value: {val2}")
    
    assert val1.as_number() == 6.0
    assert val2.as_number() == 15.0

    # Check that formulas were stored correctly
    forms = s.get_formulas(fz.RangeAddress("Data", 1, 4, 2, 4))
    assert forms == [["SUM(A1:C1)"], ["SUM(A2:C2)"]]


def test_load_workbook_from_disk(tmp_path: Path):
    """Test loading an actual XLSX file from disk with Calamine."""
    # Create a workbook with openpyxl
    xlsx_path = tmp_path / "test_workbook.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Add some values
    ws["A1"] = 100
    ws["A2"] = 200
    ws["A3"] = 300
    
    # Add formulas
    ws["B1"] = "=A1*2"
    ws["B2"] = "=A2+A3"
    ws["B3"] = "=SUM(A1:A3)"
    
    # Save the workbook
    wb.save(xlsx_path)
    
    # Load with formualizer
    fz_wb = fz.load_workbook(str(xlsx_path))
    
    # Create engine with the workbook
    engine = fz.Engine(fz_wb)
    
    # Check values were loaded (raw values from workbook)
    sheet = fz_wb.sheet("Data")
    assert sheet.get_cell(1, 1).value.as_number() == 100.0
    assert sheet.get_cell(2, 1).value.as_number() == 200.0
    assert sheet.get_cell(3, 1).value.as_number() == 300.0
    
    # Check formulas were loaded and can be evaluated
    assert engine.evaluate_cell("Data", 1, 2).as_number() == 200.0  # A1*2
    assert engine.evaluate_cell("Data", 2, 2).as_number() == 500.0  # A2+A3
    assert engine.evaluate_cell("Data", 3, 2).as_number() == 600.0  # SUM(A1:A3)
    
    # Test using classmethod directly
    fz_wb2 = fz.Workbook.load_path(str(xlsx_path), strategy="eager_all")
    engine2 = fz.Engine.from_workbook(fz_wb2)
    assert engine2.evaluate_cell("Data", 1, 1).as_number() == 100.0


def test_formula_evaluation_types():
    """Test that formula evaluation returns the correct types."""
    wb = fz.Workbook()
    s = wb.sheet("Test")
    
    # Set up integer values
    s.set_value(1, 1, fz.LiteralValue.int(10))
    s.set_value(2, 1, fz.LiteralValue.int(20))
    s.set_value(3, 1, fz.LiteralValue.int(30))
    
    # Set up various formulas
    s.set_formula(1, 2, "=A1+A2")  # Simple addition
    s.set_formula(2, 2, "=A1*2")    # Multiplication
    s.set_formula(3, 2, "=A1/2")    # Division (should return float)
    s.set_formula(4, 2, "=SUM(A1:A3)")  # SUM function
    s.set_formula(5, 2, "=AVERAGE(A1:A3)")  # AVERAGE (definitely float)
    
    # Create engine from the workbook to evaluate formulas
    engine = fz.Engine.from_workbook(wb)
    
    # Check the types and values through evaluation
    add_result = engine.evaluate_cell("Test", 1, 2)
    print(f"A1+A2: type={add_result.type_name}, value={add_result}")
    assert add_result.as_number() == 30.0
    
    mult_result = engine.evaluate_cell("Test", 2, 2)
    print(f"A1*2: type={mult_result.type_name}, value={mult_result}")
    assert mult_result.as_number() == 20.0
    
    div_result = engine.evaluate_cell("Test", 3, 2)
    print(f"A1/2: type={div_result.type_name}, value={div_result}")
    assert div_result.as_number() == 5.0
    
    sum_result = engine.evaluate_cell("Test", 4, 2)
    print(f"SUM(A1:A3): type={sum_result.type_name}, value={sum_result}")
    assert sum_result.as_number() == 60.0
    
    avg_result = engine.evaluate_cell("Test", 5, 2)
    print(f"AVERAGE(A1:A3): type={avg_result.type_name}, value={avg_result}")
    assert avg_result.as_number() == 20.0

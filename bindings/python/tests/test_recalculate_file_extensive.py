from __future__ import annotations

import shutil
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - allow skipping if not present in dev env
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CONTENT_TYPES = "http://schemas.openxmlformats.org/package/2006/content-types"
CALC_CHAIN_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/calcChain"
)


def _fixture(name: str) -> Path:
    return Path(__file__).with_name("fixtures") / name


def _copy_fixture(tmp_path: Path, name: str) -> Path:
    dest = tmp_path / name
    shutil.copy2(_fixture(name), dest)
    return dest


def _inject_calc_chain(path: Path) -> None:
    with zipfile.ZipFile(path, "r") as zin:
        entries = {info.filename: zin.read(info.filename) for info in zin.infolist()}

    rels_name = "xl/_rels/workbook.xml.rels"
    rels_root = ET.fromstring(entries[rels_name])
    rel = ET.SubElement(rels_root, f"{{{NS_PKG_REL}}}Relationship")
    rel.set("Id", "rId999")
    rel.set("Type", CALC_CHAIN_REL_TYPE)
    rel.set("Target", "calcChain.xml")
    entries[rels_name] = ET.tostring(rels_root, encoding="utf-8", xml_declaration=True)

    ct_name = "[Content_Types].xml"
    ct_root = ET.fromstring(entries[ct_name])
    override = ET.SubElement(ct_root, f"{{{NS_CONTENT_TYPES}}}Override")
    override.set("PartName", "/xl/calcChain.xml")
    override.set(
        "ContentType",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml",
    )
    entries[ct_name] = ET.tostring(ct_root, encoding="utf-8", xml_declaration=True)

    entries[
        "xl/calcChain.xml"
    ] = (
        b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        b'<calcChain xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        b'<c r="B1" i="1"/>'
        b"</calcChain>"
    )

    with zipfile.ZipFile(path, "w") as zout:
        for name, data in entries.items():
            zout.writestr(name, data)


def test_formulae_fixture_recalculate_writes_expected_cache(tmp_path: Path):
    path = _copy_fixture(tmp_path, "formulae.xlsx")
    result = fz.recalculate_file(path)

    assert result["evaluated"] == 10
    assert result["errors"] == 2
    assert result["sheets"]["Sheet1"] == {"evaluated": 10, "errors": 2}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    assert ws["A3"].value == 12345
    assert ws["A4"].value == 24690
    assert ws["A5"].value == 49380
    assert ws["A16"].value == "Düsseldorf"
    assert ws["B7"].value == "#VALUE!"
    assert ws["C10"].value == "#VALUE!"


def test_name_with_value_bug_fixture_maps_name_errors(tmp_path: Path):
    path = _copy_fixture(tmp_path, "NameWithValueBug.xlsx")
    result = fz.recalculate_file(path)

    assert result["evaluated"] == 2
    assert result["errors"] == 2
    assert result["sheets"]["Sheet1"] == {"evaluated": 2, "errors": 2}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    assert ws["A2"].value == "#NAME?"
    assert ws["A3"].value == "#NAME?"


def test_contains_chartsheets_fixture_reports_loader_error(tmp_path: Path):
    path = _copy_fixture(tmp_path, "contains_chartsheets.xlsx")
    with pytest.raises(OSError, match="load failed"):
        fz.recalculate_file(path)


def test_recalculate_removes_stale_calc_chain(tmp_path: Path):
    path = tmp_path / "calc_chain.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 2
    ws["A2"] = 3
    ws["B1"] = "=A1+A2"
    wb.save(path)

    _inject_calc_chain(path)
    with zipfile.ZipFile(path, "r") as before:
        assert "xl/calcChain.xml" in before.namelist()

    result = fz.recalculate_file(path)
    assert result["evaluated"] == 1
    assert result["errors"] == 0

    with zipfile.ZipFile(path, "r") as after:
        names = set(after.namelist())
        assert "xl/calcChain.xml" not in names

        rels = ET.fromstring(after.read("xl/_rels/workbook.xml.rels"))
        assert not rels.findall(
            f"{{{NS_PKG_REL}}}Relationship[@Type='{CALC_CHAIN_REL_TYPE}']"
        )

        ct = ET.fromstring(after.read("[Content_Types].xml"))
        assert not ct.findall(
            f"{{{NS_CONTENT_TYPES}}}Override[@PartName='/xl/calcChain.xml']"
        )

    data_only = openpyxl.load_workbook(path, data_only=True)
    assert data_only["Sheet1"]["B1"].value == 5


def test_recalculate_keeps_worksheet_formula_nodes_for_shared_like_cells(tmp_path: Path):
    path = tmp_path / "shared_like.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = 3
    ws["B1"] = "=A1*2"
    ws["B2"] = "=A2*2"
    ws["B3"] = "=A3*2"
    wb.save(path)

    result = fz.recalculate_file(path)
    assert result["evaluated"] == 3
    assert result["errors"] == 0

    with zipfile.ZipFile(path, "r") as zf:
        sheet = ET.fromstring(zf.read("xl/worksheets/sheet1.xml"))
        for ref, expected in [("B1", "2"), ("B2", "4"), ("B3", "6")]:
            node = sheet.find(f".//{{{NS_MAIN}}}c[@r='{ref}']")
            assert node is not None
            formula = node.find(f"{{{NS_MAIN}}}f")
            value = node.find(f"{{{NS_MAIN}}}v")
            assert formula is not None
            assert value is not None
            assert value.text == expected


def test_recalculate_handles_many_formulas_in_one_sheet(tmp_path: Path):
    path = tmp_path / "many_formulas.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Seed values in A1:A40.
    for row in range(1, 41):
        ws.cell(row=row, column=1).value = row

    # Fill B1:U40 with formulas that reference column A.
    # 40 rows * 20 cols = 800 formulas.
    for row in range(1, 41):
        for col in range(2, 22):
            ws.cell(row=row, column=col).value = f"=$A{row}*{col - 1}"
    wb.save(path)

    result = fz.recalculate_file(path)
    assert result["evaluated"] == 800
    assert result["errors"] == 0
    assert result["sheets"]["Sheet1"] == {"evaluated": 800, "errors": 0}

    data_only = openpyxl.load_workbook(path, data_only=True)["Sheet1"]
    assert data_only["B1"].value == 1
    assert data_only["U1"].value == 20
    assert data_only["B40"].value == 40
    assert data_only["U40"].value == 800

from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None


pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


def test_excel_1900_date_serial_no_off_by_one(tmp_path: Path):
    p = tmp_path / "date-serial.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Target: H15 (row=15, col=8)
    ws["H15"] = dt.date(2023, 3, 1)
    ws["H15"].number_format = "m/d/yyyy"

    wb.save(p)

    fz_wb = fz.load_workbook(str(p), strategy="eager_all")
    got = fz_wb.evaluate_cell("Sheet1", 15, 8)
    assert got == dt.date(2023, 3, 1)

from __future__ import annotations

import datetime as dt
from pathlib import Path

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None


pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


def test_openpyxl_date_arithmetic_propagates_date_tag(tmp_path: Path):
    p = tmp_path / "date-arithmetic.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Set C107 to a date and add a formula for C109 that does date arithmetic.
    ws["C107"] = dt.date(2024, 10, 18)
    ws["C107"].number_format = "m/d/yyyy"
    ws["C108"] = 1
    ws["C109"] = "=C107+(ROUND(C108,0)*14)"
    ws["C109"].number_format = "m/d/yyyy"

    wb.save(p)

    fz_wb = fz.load_workbook(str(p), strategy="eager_all")
    got = fz_wb.evaluate_cell("Sheet1", 109, 3)
    if isinstance(got, dt.datetime):
        got = got.date()
    assert got == dt.date(2024, 11, 1)

from io import BytesIO
from pathlib import Path

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - allow skipping if not present in dev env
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


def test_load_workbook_bytes_uses_umya_for_xlsx(xlsx_builder):
    path = xlsx_builder(lambda wb: _populate_input_workbook(wb))
    payload = Path(path).read_bytes()

    wb = fz.load_workbook_bytes(payload)
    assert wb.evaluate_cell("Sheet1", 1, 2) == 42.0

    wb2 = fz.Workbook.from_bytes(payload, backend="umya")
    assert wb2.evaluate_cell("Sheet1", 1, 2) == 42.0


def test_load_workbook_bytes_rejects_calamine_for_now(xlsx_builder):
    path = xlsx_builder(lambda wb: _populate_input_workbook(wb))
    payload = Path(path).read_bytes()

    with pytest.raises(NotImplementedError, match="calamine"):
        fz.load_workbook_bytes(payload, backend="calamine")

    with pytest.raises(NotImplementedError, match="calamine"):
        fz.Workbook.from_bytes(payload, backend="calamine")


def test_to_xlsx_bytes_roundtrip_preserves_workbook_content():
    wb = fz.Workbook()
    wb.add_sheet("Export")
    wb.set_value("Export", 1, 1, 14)
    wb.set_value("Export", 2, 1, 28)
    wb.set_formula("Export", 1, 2, "=SUM(A1:A2)")
    wb.evaluate_all()

    payload = wb.to_xlsx_bytes()
    assert isinstance(payload, bytes)
    assert len(payload) > 100

    openpyxl_wb = openpyxl.load_workbook(BytesIO(payload), data_only=False)
    ws = openpyxl_wb["Export"]
    assert ws["A1"].value == 14
    assert ws["A2"].value == 28
    assert ws["B1"].value == "=SUM(A1:A2)"

    reopened = fz.load_workbook_bytes(payload)
    assert reopened.evaluate_cell("Export", 1, 2) == 42.0


def _populate_input_workbook(wb) -> None:
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 20
    ws["A2"] = 22
    ws["B1"] = "=SUM(A1:A2)"

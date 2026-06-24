from __future__ import annotations

import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - allow skipping if not present in dev env
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _sheet_xml_by_name(path: Path) -> dict[str, ET.Element]:
    with zipfile.ZipFile(path, "r") as zf:
        workbook_root = ET.fromstring(zf.read("xl/workbook.xml"))
        rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))

        rels = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels_root.findall(
                "{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
            )
            if rel.attrib.get("Type", "").endswith("/worksheet")
        }

        out: dict[str, ET.Element] = {}
        for sheet in workbook_root.findall(f".//{{{NS_MAIN}}}sheet"):
            sheet_name = sheet.attrib["name"]
            rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            target = rels[rel_id]
            if target.startswith("/"):
                part = target.lstrip("/")
            else:
                part = str(Path("xl") / target)
            part = part.replace("\\", "/")
            out[sheet_name] = ET.fromstring(zf.read(part))
        return out


def _cell(root: ET.Element, ref: str) -> ET.Element:
    node = root.find(f".//{{{NS_MAIN}}}c[@r='{ref}']")
    assert node is not None, f"cell {ref} not found"
    return node


def _cell_payload(root: ET.Element, ref: str) -> tuple[str | None, str | None]:
    node = _cell(root, ref)
    value = node.find(f"{{{NS_MAIN}}}v")
    return node.attrib.get("t"), value.text if value is not None else None


def test_recalculate_file_in_place_writes_cached_values(tmp_path: Path):
    path = tmp_path / "in_place.xlsx"
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = 1
    ws1["A2"] = 2
    ws1["A3"] = 3
    ws1["B1"] = "=SUM(A1:A3)"
    ws1["B2"] = "=AVERAGE(A1:A3)"
    ws1["B3"] = '=IF(B1>0,"ok","bad")'
    ws1["B4"] = "=IF(A1=1,TRUE,FALSE)"
    ws1["B5"] = "=1/0"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "=Sheet1!B1+4"
    ws2["A2"] = '=IF(Sheet1!B3="ok",1,0)'
    ws2["A3"] = "=Sheet1!B3"
    wb.save(path)

    result = fz.recalculate_file(path)
    assert result["status"] == "errors_found"
    assert result["evaluated"] == 8
    assert result["errors"] == 1
    assert result["total_formulas"] == 8
    assert result["total_errors"] == 1
    assert result["sheets"]["Sheet1"] == {"evaluated": 5, "errors": 1}
    assert result["sheets"]["Sheet2"] == {"evaluated": 3, "errors": 0}
    assert result["error_summary"]["#DIV/0!"] == {
        "count": 1,
        "locations": ["Sheet1!B5"],
    }

    sheets = _sheet_xml_by_name(path)
    assert _cell_payload(sheets["Sheet1"], "B1") == (None, "6")
    assert _cell_payload(sheets["Sheet1"], "B2") == (None, "2")
    assert _cell_payload(sheets["Sheet1"], "B3") == ("str", "ok")
    assert _cell_payload(sheets["Sheet1"], "B4") == ("b", "1")
    assert _cell_payload(sheets["Sheet1"], "B5") == ("e", "#DIV/0!")
    assert _cell_payload(sheets["Sheet2"], "A1") == (None, "10")
    assert _cell_payload(sheets["Sheet2"], "A2") == (None, "1")
    assert _cell_payload(sheets["Sheet2"], "A3") == ("str", "ok")

    data_only = openpyxl.load_workbook(path, data_only=True)
    assert data_only["Sheet1"]["B1"].value == 6
    assert data_only["Sheet1"]["B2"].value == 2
    assert data_only["Sheet1"]["B3"].value == "ok"
    assert data_only["Sheet1"]["B4"].value is True
    assert data_only["Sheet1"]["B5"].value == "#DIV/0!"
    assert data_only["Sheet2"]["A1"].value == 10
    assert data_only["Sheet2"]["A2"].value == 1
    assert data_only["Sheet2"]["A3"].value == "ok"


def test_recalculate_file_output_writes_to_new_path(tmp_path: Path):
    in_path = tmp_path / "input.xlsx"
    out_path = tmp_path / "output.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 5
    ws["A2"] = 7
    ws["B1"] = "=SUM(A1:A2)"
    wb.save(in_path)

    before = openpyxl.load_workbook(in_path, data_only=True)
    assert before["Sheet1"]["B1"].value is None

    result = fz.recalculate_file(in_path, output=out_path)
    assert result["status"] == "success"
    assert result["evaluated"] == 1
    assert result["errors"] == 0
    assert result["total_formulas"] == 1
    assert result["total_errors"] == 0
    assert "error_summary" not in result

    in_after = openpyxl.load_workbook(in_path, data_only=True)
    out_after = openpyxl.load_workbook(out_path, data_only=True)
    assert in_after["Sheet1"]["B1"].value is None
    assert out_after["Sheet1"]["B1"].value == 12


def test_recalculate_file_xlfn_retry_keeps_formula_text(tmp_path: Path):
    path = tmp_path / "xlfn.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = 3
    ws["B1"] = "=_xlfn.SUM(A1:A3)"
    wb.save(path)

    result = fz.recalculate_file(path)
    assert result["status"] == "success"
    assert result["evaluated"] == 1
    assert result["errors"] == 0
    assert result["total_formulas"] == 1
    assert result["total_errors"] == 0

    sheets = _sheet_xml_by_name(path)
    assert _cell_payload(sheets["Sheet1"], "B1") == (None, "6")

    formula_node = _cell(sheets["Sheet1"], "B1").find(f"{{{NS_MAIN}}}f")
    assert formula_node is not None
    assert formula_node.text == "_xlfn.SUM(A1:A3)"

    data_only = openpyxl.load_workbook(path, data_only=True)
    with_formula = openpyxl.load_workbook(path, data_only=False)
    assert data_only["Sheet1"]["B1"].value == 6
    assert with_formula["Sheet1"]["B1"].value == "=_xlfn.SUM(A1:A3)"


def test_recalculate_file_xlfn_retry_handles_dependents(tmp_path: Path):
    path = tmp_path / "xlfn_dependent.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 1
    ws["A2"] = 2
    ws["A3"] = 3
    ws["B1"] = "=_xlfn.SUM(A1:A3)"
    ws["C1"] = "=B1+1"
    wb.save(path)

    result = fz.recalculate_file(path)
    assert result["status"] == "success"
    assert result["evaluated"] == 2
    assert result["errors"] == 0

    data_only = openpyxl.load_workbook(path, data_only=True)
    assert data_only["Sheet1"]["B1"].value == 6
    assert data_only["Sheet1"]["C1"].value == 7


def test_workbook_recalculate_formula_cells_batches_targets(tmp_path: Path):
    path = tmp_path / "batch_targets.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 4
    ws["A2"] = 9
    ws["B1"] = "=A1+A2"
    ws["B2"] = "=A2-A1"
    wb.save(path)

    workbook = fz.Workbook.from_path(str(path))

    assert workbook.recalculate_formula_cells(
        [("Sheet1", 1, 2), ("Sheet1", 2, 2)]
    ) == [13, 5]


def test_recalculate_file_maps_name_errors(tmp_path: Path):
    path = tmp_path / "name_error.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "=NOPE(1)"
    wb.save(path)

    result = fz.recalculate_file(path)
    assert result == {
        "status": "errors_found",
        "evaluated": 1,
        "errors": 1,
        "total_formulas": 1,
        "total_errors": 1,
        "sheets": {"Sheet1": {"evaluated": 1, "errors": 1}},
        "error_summary": {"#NAME?": {"count": 1, "locations": ["Sheet1!A1"]}},
    }

    sheets = _sheet_xml_by_name(path)
    assert _cell_payload(sheets["Sheet1"], "A1") == ("e", "#NAME?")

    data_only = openpyxl.load_workbook(path, data_only=True)
    assert data_only["Sheet1"]["A1"].value == "#NAME?"


def test_recalculate_file_caps_error_locations_per_type(tmp_path: Path):
    path = tmp_path / "error_cap.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 26):
        ws[f"A{row}"] = "=1/0"
    wb.save(path)

    result = fz.recalculate_file(path)
    assert result["status"] == "errors_found"
    assert result["evaluated"] == 25
    assert result["errors"] == 25
    assert result["total_formulas"] == 25
    assert result["total_errors"] == 25

    error_info = result["error_summary"]["#DIV/0!"]
    assert error_info["count"] == 25
    assert len(error_info["locations"]) == 20
    assert error_info["locations"][0] == "Sheet1!A1"
    assert error_info["locations"][-1] == "Sheet1!A20"
    assert error_info["locations_truncated"] == 5

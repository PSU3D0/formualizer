from pathlib import Path

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - allow skipping if not present in dev env
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


def test_index_supports_single_index_on_one_row_horizontal_range(tmp_path: Path):
    path = tmp_path / "index-horizontal.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 10
    ws["B1"] = 20
    ws["C1"] = 30
    ws["A2"] = "=INDEX(A1:C1,1)"
    ws["B2"] = "=INDEX(A1:C1,2)"
    ws["C2"] = "=INDEX(A1:C1,3)"
    ws["A3"] = "=INDEX(A1:C1,MATCH(20,A1:C1,0))"
    wb.save(path)

    workbook = fz.load_workbook(str(path), strategy="eager_all")

    assert workbook.evaluate_cell("Sheet1", 2, 1) == 10.0
    assert workbook.evaluate_cell("Sheet1", 2, 2) == 20.0
    assert workbook.evaluate_cell("Sheet1", 2, 3) == 30.0
    assert workbook.evaluate_cell("Sheet1", 3, 1) == 20.0

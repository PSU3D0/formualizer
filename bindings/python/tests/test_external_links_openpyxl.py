"""In-workbook external references (spec §10): `[1]Sheet1!A1`.

Excel stores cross-workbook references inside the same file via
`xl/externalLinks/externalLinkN.xml` parts that cache the referenced values
(`<sheetData>`); Excel never recalculates external links and surfaces those
cached values. formualizer must do the same instead of failing with
`#NAME?: Undefined name: [1]Sheet1!A1`.
"""

import zipfile

import pytest

import formualizer as fz

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - allow skipping if not present in dev env
    openpyxl = None

pytestmark = pytest.mark.skipif(openpyxl is None, reason="openpyxl not installed")


def build_inbook_external_link_xlsx(tmp_path, sheet_data="42", with_cache=True):
    """Build an `.xlsx` with `=[1]Sheet1!A1` and an external link part caching A1."""
    p = tmp_path / "external_links.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 42
    wb.save(p)

    NS = "http://schemas.openxmlformats.org/package/2006/relationships"
    SML = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    with zipfile.ZipFile(p, "r") as z:
        content_types = z.read("[Content_Types].xml").decode("utf-8")
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        wbrels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        sheet1 = z.read("xl/worksheets/sheet1.xml").decode("utf-8")

    # Always add the formula cell so the reference exists even without a cache.
    sheet1 = sheet1.replace(
        "</row>",
        f'      <c r="B1">\n        <f t="shared" ref="B1" si="0">[1]Sheet1!A1</f>\n'
        f"        <v>{sheet_data}</v>\n      </c>\n    </row>",
    )
    if with_cache:
        content_types = content_types.replace(
            '<Override PartName="/xl/styles.xml"',
            '<Override PartName="/xl/externalLinks/externalLink1.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"/>\n'
            '  <Override PartName="/xl/styles.xml"',
        )
        wbxml = wbxml.replace(
            "</sheets>",
            '</sheets>\n  <externalReferences>\n    <externalReference r:id="rId4"/>\n  </externalReferences>',
        )
        wbrels = wbrels.replace(
            "</Relationships>",
            f'  <Relationship Id="rId4" Type="{NS}/externalLink" '
            f'Target="externalLinks/externalLink1.xml"/>\n</Relationships>',
        )
        extlink = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<externalLink xmlns="{SML}">
  <externalBook xmlns:r="{NS}" r:id="rId1">
    <sheetNames>
      <sheetName val="Sheet1"/>
    </sheetNames>
    <sheetDataSet>
      <sheetData sheetId="0">
        <row r="1">
          <c r="A1"><v>{sheet_data}</v></c>
        </row>
      </sheetData>
    </sheetDataSet>
  </externalBook>
</externalLink>"""
        extlink_rels = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS}">
  <Relationship Id="rId1" Type="{NS}/externalBook" Target="../workbook.xml"/>
</Relationships>"""

    with zipfile.ZipFile(p.with_suffix(".tmp"), "w", zipfile.ZIP_DEFLATED) as zt:
        for name, content in (
            ("[Content_Types].xml", content_types),
            ("xl/workbook.xml", wbxml),
            ("xl/_rels/workbook.xml.rels", wbrels),
            ("xl/worksheets/sheet1.xml", sheet1),
        ):
            zt.writestr(name, content)
        if with_cache:
            zt.writestr("xl/externalLinks/externalLink1.xml", extlink)
            zt.writestr("xl/externalLinks/_rels/externalLink1.xml.rels", extlink_rels)
        with zipfile.ZipFile(p, "r") as zo:
            for name in zo.namelist():
                if name not in (
                    "[Content_Types].xml",
                    "xl/workbook.xml",
                    "xl/_rels/workbook.xml.rels",
                    "xl/worksheets/sheet1.xml",
                ):
                    zt.writestr(name, zo.read(name))
    p.with_suffix(".tmp").replace(p)
    return p


def test_inbook_external_reference_evaluates_from_cached_values(tmp_path):
    p = build_inbook_external_link_xlsx(tmp_path)
    wb = fz.load_workbook(str(p))
    wb.evaluate_all()
    assert wb.get_value("Sheet1", 1, 2) == 42.0
    assert wb.get_formula("Sheet1", 1, 2) == "=[1]Sheet1!A1"


def test_inbook_external_reference_without_cache_fails_gracefully(tmp_path):
    p = build_inbook_external_link_xlsx(tmp_path, with_cache=False)
    wb = fz.load_workbook(str(p))
    with pytest.raises(fz.ExcelEvaluationError):
        wb.evaluate_all()
